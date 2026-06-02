from pathlib import Path

from openpyxl import load_workbook

RUTA_EXCEL = Path(__file__).resolve().parent.parent / "productos.xlsx"


def obtener_todos_los_productos():
    """Lee todos los productos del archivo Excel y devuelve una lista de diccionarios."""
    try:
        libro = load_workbook(RUTA_EXCEL)
    except FileNotFoundError:
        return []

    hoja = libro.active
    productos = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if all(valor is None for valor in fila):
            continue

        id_producto, nombre, precio, cantidad = fila[:4]
        productos.append(
            {
                "Id": id_producto,
                "Nombre": nombre,
                "Precio": precio,
                "Cantidad": cantidad,
            }
        )

    libro.close()
    return productos


def obtener_producto_por_id(id_buscado):
    """Busca un producto por Id en el archivo Excel y devuelve su diccionario o None."""
    try:
        libro = load_workbook(RUTA_EXCEL)
    except FileNotFoundError:
        return None

    hoja = libro.active

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if all(valor is None for valor in fila):
            continue

        id_producto, nombre, precio, cantidad = fila[:4]
        if id_producto == id_buscado:
            libro.close()
            return {
                "Id": id_producto,
                "Nombre": nombre,
                "Precio": precio,
                "Cantidad": cantidad,
            }

    libro.close()
    return None


def crear_producto(id_producto, nombre, precio, cantidad):
    """Crea un nuevo producto en el archivo Excel y devuelve True si fue exitoso."""
    try:
        libro = load_workbook(RUTA_EXCEL)
    except FileNotFoundError:
        return False

    hoja = libro.active

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if all(valor is None for valor in fila):
            continue

        id_existente = fila[0]
        if id_existente == id_producto:
            libro.close()
            return False

    hoja.append([id_producto, nombre, precio, cantidad])
    libro.save(RUTA_EXCEL)
    libro.close()
    return True


def actualizar_producto(id_producto, nombre, precio, cantidad):
    """Actualiza un producto en el archivo Excel y devuelve True si fue exitoso."""
    try:
        libro = load_workbook(RUTA_EXCEL)
    except FileNotFoundError:
        return False

    hoja = libro.active

    for indice_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        if all(valor is None for valor in fila):
            continue

        id_existente = fila[0]
        if id_existente == id_producto:
            hoja.cell(row=indice_fila, column=2, value=nombre)
            hoja.cell(row=indice_fila, column=3, value=precio)
            hoja.cell(row=indice_fila, column=4, value=cantidad)
            libro.save(RUTA_EXCEL)
            libro.close()
            return True

    libro.close()
    return False


def eliminar_producto(id_producto):
    """Elimina un producto del archivo Excel y devuelve True si fue exitoso."""
    try:
        libro = load_workbook(RUTA_EXCEL)
    except FileNotFoundError:
        return False

    hoja = libro.active

    for indice_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        if all(valor is None for valor in fila):
            continue

        id_existente = fila[0]
        if id_existente == id_producto:
            hoja.delete_rows(indice_fila, 1)
            libro.save(RUTA_EXCEL)
            libro.close()
            return True

    libro.close()
    return False
